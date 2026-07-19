#!/usr/bin/env python3
import argparse
import json
from collections import Counter
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


CONIFER_SHEETS = ["Spruce", "Fir", "Pine", "Other Ever."]
BROADLEAF_SHEETS = [
    "Shrubs-1",
    "Shrubs-2",
    "Small Trees-1",
    "Small Trees-2",
    "Large Trees-1",
    "Large Trees-2",
]
BIGGER_STUFF_SHEETS = ["Bigger Stuff 1", "Bigger Stuff 2"]
DEFAULT_DESCRIPTION = ""


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def money(value, divisor=1):
    if value in (None, ""):
        return ""

    if isinstance(value, str):
        value = value.strip().replace("$", "").replace(",", "")
        if not value or value in {"*", "`"}:
            return ""

    try:
        number = Decimal(str(value)) / Decimal(str(divisor))
    except Exception:
        return ""

    return str(number.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP).normalize())


class CatalogImporter:
    def __init__(self):
        self.rows = []
        self.next_id = 1
        self.skipped = Counter()

    def add_row(self, name, latin, size="", age="", prices=None, category=""):
        prices = prices or {}
        name = clean(name)
        latin = clean(latin)

        if not name or not latin:
            self.skipped["missing name or latin"] += 1
            return

        row = {
            "available": "Yes",
            "id": str(self.next_id),
            "name": name,
            "latin": latin,
            "description": DEFAULT_DESCRIPTION,
        }

        if size:
            row["size"] = clean(size)
        if age:
            row["age"] = clean(age)

        for key in [
            "fivehundredprice",
            "onehundredprice",
            "twentyfiveprice",
            "oneprice",
        ]:
            if prices.get(key):
                row[key] = prices[key]

        if category:
            row["type"] = category

        self.rows.append(row)
        self.next_id += 1

    def import_conifer_sheet(self, worksheet):
        current_name = ""
        current_latin = ""
        pending_rows = []
        price_start = 8 if worksheet.title == "Spruce" else 7

        def flush_pending_rows():
            while pending_rows:
                item = pending_rows.pop(0)
                self.add_row(
                    current_name,
                    current_latin,
                    item["size"],
                    item["age"],
                    item["prices"],
                    worksheet.title,
                )

        for values in worksheet.iter_rows(min_row=7, values_only=True):
            raw_first = values[0] if len(values) > 0 else None
            first = clean(raw_first)
            indented = isinstance(raw_first, str) and raw_first.startswith(" ")
            size = clean(values[4] if len(values) > 4 else "")
            age = clean(values[5] if len(values) > 5 else "")
            prices = {
                "twentyfiveprice": money(
                    values[price_start] if len(values) > price_start else None
                ),
                "onehundredprice": money(
                    values[price_start + 1] if len(values) > price_start + 1 else None,
                    100,
                ),
                "fivehundredprice": money(
                    values[price_start + 2] if len(values) > price_start + 2 else None,
                    1000,
                ),
            }

            if first and not indented:
                if current_name and current_latin:
                    flush_pending_rows()
                current_name = first
                current_latin = ""
            elif first and indented and not current_latin:
                current_latin = first
                flush_pending_rows()

            if size and any(prices.values()):
                pending_rows.append({"size": size, "age": age, "prices": prices})

        if current_name and current_latin:
            flush_pending_rows()

    def import_size_price_matrix_sheet(self, worksheet, price_key):
        size_headers = [clean(cell.value) for cell in worksheet[5]]

        for values in worksheet.iter_rows(min_row=6, values_only=True):
            name = clean(values[0] if len(values) > 0 else "")
            latin = clean(values[2] if len(values) > 2 else "")
            age = clean(values[4] if len(values) > 4 else "")

            if not name and not latin:
                continue

            for column_index, size in enumerate(size_headers):
                if not size or column_index >= len(values):
                    continue

                price = money(values[column_index])
                if price:
                    self.add_row(
                        name,
                        latin,
                        size,
                        age if price_key == "fivehundredprice" else "",
                        {price_key: price},
                        worksheet.title,
                    )

    def import_workbook(self, workbook):
        for sheet_name in CONIFER_SHEETS:
            self.import_conifer_sheet(workbook[sheet_name])

        for sheet_name in BROADLEAF_SHEETS:
            self.import_size_price_matrix_sheet(workbook[sheet_name], "fivehundredprice")

        for sheet_name in BIGGER_STUFF_SHEETS:
            self.import_size_price_matrix_sheet(workbook[sheet_name], "onehundredprice")

        return self.rows


def parse_args():
    parser = argparse.ArgumentParser(
        description="Convert Alpha Nurseries price-list Excel files into catalog data.json."
    )
    parser.add_argument("excel_file", help="Path to the Excel workbook to import.")
    parser.add_argument(
        "-o",
        "--output",
        default="data.json",
        help="JSON output path. Defaults to data.json.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and report without writing the output file.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    excel_path = Path(args.excel_file)
    output_path = Path(args.output)

    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    importer = CatalogImporter()
    rows = importer.import_workbook(workbook)

    type_counts = Counter(row.get("type", "Uncategorized") for row in rows)
    price_counts = Counter(
        price_key
        for row in rows
        for price_key in [
            "fivehundredprice",
            "onehundredprice",
            "twentyfiveprice",
            "oneprice",
        ]
        if price_key in row
    )

    print(f"Imported rows: {len(rows)}")
    print(f"Unique products: {len({row['name'] for row in rows})}")
    print("Rows by sheet:")
    for sheet_name, count in type_counts.items():
        print(f"  {sheet_name}: {count}")
    print("Price fields:")
    for price_key, count in price_counts.items():
        print(f"  {price_key}: {count}")

    if importer.skipped:
        print("Skipped rows:")
        for reason, count in importer.skipped.items():
            print(f"  {reason}: {count}")

    if args.dry_run:
        print("Dry run complete. No files written.")
        return

    output_path.write_text(json.dumps(rows, indent=2) + "\n")
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
